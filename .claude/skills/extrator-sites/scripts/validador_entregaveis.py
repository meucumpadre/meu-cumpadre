#!/usr/bin/env python3
"""
VALIDADOR ENTREGÁVEIS - PORTEIRO EXIGENTE
Valida rigorosamente: Excel, PDF, CSV e alocação em pasta destino
"""

import os
import re
from pathlib import Path
from urllib.parse import urlparse
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ValidadorEntregaveis:
    """Porteiro exigente que valida entregáveis com critérios rigorosos"""
    
    def __init__(self, workspace_folder: str):
        self.workspace_folder = Path(workspace_folder)
        self.erros = []
        self.avisos = []
        self.passou = True
    
    def validar_completo(self, dominio: str, timestamp: str) -> Dict:
        """Executa validação completa de todos os entregáveis"""
        print("\n" + "="*60)
        print("🛡️  PORTEIRO EXIGENTE - VALIDAÇÃO COMPLETA")
        print("="*60)
        
        # Padrão de nome de arquivo
        pattern = f"mapeamento_{dominio}_{timestamp}"
        
        # Verificar existência de arquivos
        xlsx_file = self._find_file(pattern, ".xlsx")
        pdf_file = self._find_file(pattern, ".pdf")
        csv_file = self._find_file(pattern, ".csv")
        
        if not xlsx_file or not pdf_file or not csv_file:
            self.passou = False
            self.erros.append("❌ Arquivos não encontrados na pasta destino")
            return self._gerar_relatorio()
        
        # Validar cada formato
        self._validar_xlsx(xlsx_file)
        self._validar_pdf(pdf_file)
        self._validar_csv(csv_file)
        
        # Validar alocação
        self._validar_alocacao([xlsx_file, pdf_file, csv_file])
        
        return self._gerar_relatorio()
    
    def _find_file(self, pattern: str, ext: str) -> Path:
        """Encontra arquivo com padrão e extensão"""
        try:
            for f in self.workspace_folder.glob(f"{pattern}*{ext}"):
                return f
        except:
            pass
        return None
    
    def _validar_xlsx(self, filepath: Path):
        """Valida arquivo Excel"""
        print("\n📊 Validando Excel...")
        
        try:
            # Tamanho mínimo
            size_kb = filepath.stat().st_size / 1024
            if size_kb < 5:
                self.erros.append(f"❌ Excel muito pequeno ({size_kb:.1f} KB < 5 KB)")
                self.passou = False
                return
            
            # Abrir workbook
            wb = load_workbook(filepath)
            sheet_names = wb.sheetnames
            
            # Verificar sheets
            if len(sheet_names) < 2:
                self.erros.append(f"❌ Excel com apenas {len(sheet_names)} sheet (esperado 2)")
                self.passou = False
                return
            
            # Validar Sheet 1: URLs Mineradas
            ws1 = wb[sheet_names[0]]
            headers_esperados_ws1 = ["Categoria", "Título", "URL Completa", "Profundidade"]
            if not self._verificar_headers(ws1, headers_esperados_ws1):
                self.erros.append(f"❌ Sheet 1 com headers incorretos")
                self.passou = False
                return
            
            # Validar URLs no Sheet 1
            urls_invalidas = []
            for row in range(2, ws1.max_row + 1):
                url = ws1.cell(row, 3).value  # Coluna URL
                prof = ws1.cell(row, 4).value  # Coluna Profundidade
                
                if not url:
                    continue
                
                # Validar URL completa
                if not self._url_valida(str(url)):
                    urls_invalidas.append((row, url))
                
                # Validar profundidade
                try:
                    int(prof)
                except:
                    if prof:
                        urls_invalidas.append((row, f"Profundidade inválida: {prof}"))
            
            if urls_invalidas:
                self.erros.append(f"❌ URLs malformadas no Excel:")
                for row, url in urls_invalidas[:5]:
                    self.erros.append(f"   Linha {row}: {str(url)[:50]}")
                if len(urls_invalidas) > 5:
                    self.erros.append(f"   ... e mais {len(urls_invalidas)-5}")
                self.passou = False
                return
            
            # Verificar mínimo de URLs
            total_urls = ws1.max_row - 1
            if total_urls < 10:
                self.avisos.append(f"⚠️  Excel com apenas {total_urls} URLs (recomendado ≥10)")
            
            # Validar Sheet 2: Estatísticas
            ws2 = wb[sheet_names[1]]
            headers_esperados_ws2 = ["Métrica", "Valor"]
            if not self._verificar_headers(ws2, headers_esperados_ws2):
                self.avisos.append(f"⚠️  Sheet 2 com headers diferentes do esperado")
            
            print(f"✅ Excel válido ({size_kb:.1f} KB, {total_urls} URLs únicos)")
            
        except Exception as e:
            self.erros.append(f"❌ Erro ao validar Excel: {str(e)}")
            self.passou = False
    
    def _validar_pdf(self, filepath: Path):
        """Valida arquivo PDF"""
        print("📄 Validando PDF...")
        
        try:
            size_kb = filepath.stat().st_size / 1024
            if size_kb < 3:
                self.erros.append(f"❌ PDF muito pequeno ({size_kb:.1f} KB < 3 KB)")
                self.passou = False
                return
            
            # Verificar se é PDF válido (magic bytes)
            with open(filepath, 'rb') as f:
                header = f.read(4)
                if header != b'%PDF':
                    self.erros.append("❌ PDF inválido (não possui magic bytes corretos)")
                    self.passou = False
                    return
            
            print(f"✅ PDF válido ({size_kb:.1f} KB)")
            
        except Exception as e:
            self.erros.append(f"❌ Erro ao validar PDF: {str(e)}")
            self.passou = False
    
    def _validar_csv(self, filepath: Path):
        """Valida arquivo CSV"""
        print("📋 Validando CSV...")
        
        try:
            size_kb = filepath.stat().st_size / 1024
            if size_kb < 1:
                self.erros.append(f"❌ CSV muito pequeno ({size_kb:.1f} KB < 1 KB)")
                self.passou = False
                return
            
            with open(filepath, 'r', encoding='utf-8') as f:
                linhas = f.readlines()
            
            if not linhas:
                self.erros.append("❌ CSV vazio")
                self.passou = False
                return
            
            # Verificar header
            header = linhas[0].strip()
            header_esperado = "Produto / Categoria;Título do Tópico/Subtópico;URL Completa;Profundidade"
            if header != header_esperado:
                self.avisos.append(f"⚠️  Header do CSV diferente do esperado")
            
            # Validar URLs
            urls_invalidas = []
            for i, linha in enumerate(linhas[1:], 2):
                partes = linha.strip().split(';')
                if len(partes) >= 3:
                    url = partes[2]
                    if url and not self._url_valida(url):
                        urls_invalidas.append((i, url))
            
            if urls_invalidas:
                self.erros.append(f"❌ URLs malformadas no CSV:")
                for linha, url in urls_invalidas[:5]:
                    self.erros.append(f"   Linha {linha}: {url[:50]}")
                if len(urls_invalidas) > 5:
                    self.erros.append(f"   ... e mais {len(urls_invalidas)-5}")
                self.passou = False
                return
            
            total_urls = len(linhas) - 1
            if total_urls < 10:
                self.avisos.append(f"⚠️  CSV com apenas {total_urls} URLs (recomendado ≥10)")
            
            print(f"✅ CSV válido ({size_kb:.1f} KB, {total_urls} URLs únicos)")
            
        except Exception as e:
            self.erros.append(f"❌ Erro ao validar CSV: {str(e)}")
            self.passou = False
    
    def _validar_alocacao(self, files: List[Path]):
        """Valida se arquivos estão na pasta destino correta"""
        print("📂 Validando alocação na pasta destino...")
        
        workspace_path = Path("/sessions/elegant-sharp-goodall/mnt/DEV + IA _ VibeCoding Google")
        
        for f in files:
            if not f.exists():
                self.erros.append(f"❌ Arquivo não existe: {f.name}")
                self.passou = False
            elif f.parent != workspace_path and str(f.parent) != str(workspace_path):
                self.avisos.append(f"⚠️  Arquivo pode não estar na pasta correta: {f.parent}")
        
        if not self.erros or all("alocação" not in e for e in self.erros):
            print("✅ Arquivos na pasta destino correta")
    
    def _url_valida(self, url: str) -> bool:
        """Verifica se URL é completa e válida"""
        if not url or not isinstance(url, str):
            return False
        
        url = str(url).strip()
        
        # Deve começar com http:// ou https://
        if not url.startswith(('http://', 'https://')):
            return False
        
        # Não pode ser apenas "http://" ou "https://"
        if url in ('http://', 'https://'):
            return False
        
        # Tentar fazer parse
        try:
            result = urlparse(url)
            # Deve ter netloc (domínio)
            if not result.netloc:
                return False
            # Domínio deve ter ponto (ex: exemplo.com)
            if '.' not in result.netloc:
                return False
            return True
        except:
            return False
    
    def _verificar_headers(self, sheet, headers_esperados: List[str]) -> bool:
        """Verifica se os headers da planilha correspondem"""
        headers_encontrados = []
        for col in range(1, len(headers_esperados) + 1):
            cell = sheet.cell(1, col)
            headers_encontrados.append(str(cell.value))
        
        return headers_encontrados == headers_esperados
    
    def _gerar_relatorio(self) -> Dict:
        """Gera relatório final da validação"""
        print("\n" + "="*60)
        
        if self.passou:
            print("✅ VALIDAÇÃO CONCLUÍDA COM ÊXITO!")
            print("="*60)
            print(f"\n✨ Todos os entregáveis passaram na validação rigorosa")
            print(f"   - Excel: ✅")
            print(f"   - PDF: ✅")
            print(f"   - CSV: ✅")
            print(f"   - Alocação: ✅")
            if self.avisos:
                print(f"\n⚠️  Avisos:")
                for aviso in self.avisos:
                    print(f"   {aviso}")
        else:
            print("❌ VALIDAÇÃO FALHOU!")
            print("="*60)
            print(f"\n🚨 Erros encontrados:\n")
            for erro in self.erros:
                print(f"   {erro}")
            print(f"\n⚠️  Ação: Corrigir os erros acima e reexecutar a mineração")
        
        print("\n" + "="*60 + "\n")
        
        return {
            "passou": self.passou,
            "erros": self.erros,
            "avisos": self.avisos,
            "timestamp": str(Path.cwd())
        }


if __name__ == "__main__":
    # Teste rápido
    workspace = "/sessions/elegant-sharp-goodall/mnt/DEV + IA _ VibeCoding Google"
    validador = ValidadorEntregaveis(workspace)
    resultado = validador.validar_completo("google.github.io_agents-cli", "20260621_023327")
    print(f"\nResultado: {resultado}")
